import os
from os.path import splitext, basename, isfile
import sys
import cv2

import threading
import signal
import numpy as np
import time,glob
import math
import subprocess
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
# from auto_test.Acc.execmap import associate_map
from Acc.execmap import associate_map

def get_video_dimensions(file_path):
    try:
        cap = cv2.VideoCapture(file_path)
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        return width, height
    except Exception as e:
        print(f"Error: {e}")
        return None
    finally:
        if cap.isOpened():
            cap.release()

class DIVA_measure(object):
    def __init__(self, use_DIVA, status, video_path, process_num = 1, single_stream_num = 1, all_fps = 0):
        self.width, self.height = self.get_video_dimensions(video_path)
        self.cpu = []
        self.gpu = []
        self.mem = []
        self.p_num = process_num
        self.s_num = single_stream_num
        self.all_fps = all_fps
        self.avg_fps = self.calc_avg(all_fps, process_num, single_stream_num)
        self.use_DIVA = use_DIVA
        self.status = status
        self.video_path = video_path
        self.video_name = splitext(basename(self.video_path))[0]
        self.run_error = False
    
    def calc_avg(self, all_fps, process_num, single_stream_num):
        return 0 if (process_num==0 or single_stream_num == 0) else all_fps/(process_num*single_stream_num) 
    
    def init_info(self, p_num, s_num):
        self.cpu = []
        self.gpu = []
        self.mem = []
        self.p_num = p_num
        self.s_num = s_num
        self.all_fps = 0 
        self.avg_fps = 0
        self.run_error = False

    def get_all_fps(self, fps_arr):
        for pn in range(len(fps_arr)):
            self.all_fps += fps_arr[pn]
        
        self.avg_fps = 0 if (self.p_num==0 or self.s_num == 0) else self.all_fps/(self.p_num*self.s_num) 
    
    def round(self, val):
        return round(val,2)

    def save_to_excel(self, file_name):
        execl_path = "./save_res_excel/" + file_name + ".xlsx"
        #print(execl_path)
        if isfile(execl_path):
            workbook = load_workbook(execl_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            data_name = ['name','width','height','task_case','all_stream','process_num','stream_num','cpu usage', 'gpu usage', 'mem usage', 'fps', 'avg_fps']
            for i, value in enumerate(data_name):
                sheet.cell(1,i+1,value) 

        last_row = sheet.max_row+1

        #task_case = "DIVA" if self.use_DIVA else "AI"
        data_res = [self.video_name, self.width, self.height, self.status, self.p_num * self.s_num,self.p_num, self.s_num, 
                    self.round(np.mean(self.cpu)), self.round(np.mean(self.gpu)), self.round(np.mean(self.mem)), self.round(self.all_fps),self.round(self.avg_fps)]
        for i, value in enumerate(data_res):
            #sheet.cell(last_row,i+1).number_format = numbers.FORMAT_NUMBER_00
            sheet.cell(last_row,i+1,value)  

        # Save the workbook
        workbook.save(execl_path)

    def get_video_dimensions(self, file_path):
        try:
            cap = cv2.VideoCapture(file_path)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            return width, height
        except Exception as e:
            print(f"Error: {e}")
            return None
        finally:
            if cap.isOpened():
                cap.release()

    def kill_process(self, PID):
        for pid in PID:
            process = subprocess.Popen(["kill", f"{pid}"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=False)
            # Wait for the process to complete and get the output
            stdout, stderr = process.communicate()
            exit_code = process.returncode
            print(f"kill PID : {pid}")

class DIVA_gt(object):
    
    class DS(object):
        DS_GMOT = 'GMOT'
        DS_MOT = 'MOT'
        DS_Imagenet_VID = 'Imagenet_VID'

    def __init__(self, video_path):
        self.gt_result = []
        self.gt_only_pt = []
        self.gt_number = 0
        self.gt_acc = 0.0        
        self.video_name = splitext(basename(video_path))[0]
        gt_name = video_path.split(os.sep)[-2]

        print(self.video_name)
        print(gt_name)
        self.gt_path = ""
        if gt_name == "MOT":
            # self.gt_path = f"./videos/MOT/{self.video_name}/gt/gt.txt"
            self.gt_path = f"./mAP_test/{self.video_name}.txt"
            self.MOT(self.gt_path)
        elif gt_name == "GMOT":
            # self.gt_path = f"./videos/GMOT/track_label/{self.video_name}.txt"
            self.gt_path = f"./mAP_test/{self.video_name}.txt"
            self.GMOT(self.gt_path)    
        elif gt_name == "Imagenet_VID":
            self.gt_path = f"./mAP_test/Imagenet_VID/{self.video_name}.json"
            self.Imagenet_VID(self.gt_path)   
        else:
            print("no GT")                  

        self.gt_name = gt_name        

    def clean(self):
        self.gt_result = []
        self.gt_only_pt = []
        self.gt_number = 0
        self.gt_acc = 0.0

    def GMOT(self, gt_path):
        if not isfile(gt_path):
            print(f"{gt_path} not found, no ground truth")
            return
        
        trk_label = np.loadtxt(gt_path, delimiter=',')
        self.gt_result,self.gt_only_pt = self.get_all_label(trk_label, 0)

    def MOT(self, gt_path):
        if not isfile(gt_path):
            print(f"{gt_path} not found, no ground truth")
            return

        trk_label = np.loadtxt(gt_path, delimiter=',')
        self.gt_result,self.gt_only_pt = self.get_all_label(trk_label, 1)

    def Imagenet_VID(self, gt_path):
        if not isfile(gt_path):
            print(f"{gt_path} not found, no ground truth")
            return

        json_data = self.read_json_file(gt_path)
        all_bbox = []
        for frame_info in json_data['trajectories']:
            bbox = []
            for obj_info in frame_info:
                bbox.append([obj_info['bbox']['xmin'], obj_info['bbox']['ymin'], obj_info['bbox']['xmax'], obj_info['bbox']['ymax']])
            
            all_bbox.append(bbox)
        
        self.gt_result = all_bbox
        self.gt_only_pt = all_bbox
    
    def read_json_file(self, file_path):
        try:
            with open(file_path, 'r') as file:
                data = json.load(file)
                return data
        except Exception as e:
            print(f"Error: {e}")
            return None
    
    def get_all_label(self, label, start_index):
        max_frame = 0
        for det in label:
            max_frame = max(max_frame, int(det[0]))
        
        frame_count = max_frame if start_index == 1 else max_frame + 1 ## if label start at 0 , max_frame need + 1
        all_label = [[] for _ in range(frame_count)] 
        pt_label = [[] for _ in range(frame_count)]
        for det in label:
            det[4] = det[2]+det[4] # x2 = x1 + w
            det[5] = det[3]+det[5] # y2 = y1 + h        
            all_label[int(det[0]) - start_index].append(det)
            pt_label[int(det[0]) - start_index].append(det[2:6])

        return all_label, pt_label

    def calc_acc(self, res_list, frame_num, is_async):

        if is_async:
            frame_num -= 1
        
        if frame_num < 0:
            return -1

        if len(self.gt_only_pt) <= 0 or len(self.gt_only_pt) <= frame_num:
            return -1

        gt = self.gt_only_pt[frame_num]
        # print("gt:", gt)
        if len(gt) <= 0:
            return -1
        
        acc = 1
        if len(res_list) == 0:
            acc = 0
        else:
            # res = res_list[:,1:5]
            res = np.array(res_list)
            # print("gt :", gt)
            # print("res :", res)
            matches, unmatched_detections, unmatched_trackers = associate_map(gt, res, 0.5)
            # print(matches, unmatched_detections, unmatched_trackers)
            acc = float(len(matches))/(len(matches)+len(unmatched_detections))
        
        self.gt_number += 1    
        self.gt_acc += acc    
        return acc

    def draw_gt(self, frame, frame_num, is_async):
        if is_async:
            frame_num -= 1
        
        if frame_num < 0:
            return

        if len(self.gt_only_pt) <= frame_num:
            return

        for label in self.gt_only_pt[frame_num]:
            cv2.rectangle(frame, (int(label[0]), int(label[1])), (int(label[2]), int(label[3])), (0, 255, 0), 5)


class DIVA_info(object):
    def __init__(self, width, height):
        self.Enable_calc = True
        self.Width = width
        self.Height = height
        self.Frame_count = 0
        self.AI_count = 0
        self.AI_time = 0 #ms
        self.Trk_count = 0
        self.Trk_time = 0 #ms
        self.Trk_AI_ratio = 0
        self.Obj_avgnum = 0
        self.Obj_maxnum = 0
        self.Obj_avgarea = 0
        self.Acc = 0.0
        self.FPS = 0.0

    def clean(self):
        self.Frame_count = 0
        self.AI_count = 0
        self.Width = 0
        self.Height = 0
        self.AI_time = 0 #ms
        self.Trk_count = 0
        self.Trk_time = 0 #ms
        self.Trk_AI_ratio = 0
        self.Obj_avgnum = 0
        self.Obj_maxnum = 0
        self.Obj_avgarea = 0
        self.Acc = 0.0
        self.FPS = 0.0

    def get_info(self, res_list, ait, trkt):
        if ait != 0: 
            self.AI_time += ait
            self.AI_count += 1
            
        elif trkt != 0: 
            self.Trk_time += trkt
            self.Trk_count += 1

        for res in res_list:
            # print('res :', res)
            # print((res[2]-res[0]) * (res[3]-res[1]))
            if len(res) == 4:
                self.Obj_avgarea += ((res[2]-res[0]) * (res[3]-res[1]))

        obj_number = len(res_list)
        self.Obj_avgnum += obj_number
        self.Obj_maxnum = obj_number if obj_number > self.Obj_maxnum else self.Obj_maxnum

    def calc_info(self, gt_acc, gt_len):
        #print(f"gt_acc : {gt_acc}, gt_len : {gt_len}, frame_count : {self.Frame_count}")
        # self.Trk_AI_ratio = self.round(self.Trk_count/float(self.AI_count))
        # self.AI_time  = self.round(self.AI_time/self.AI_count * 1000)
        # self.Trk_time = 0 if self.Trk_count == 0 else self.round(self.Trk_time/self.Trk_count * 1000)
        self.Obj_avgnum = self.round(self.Obj_avgnum/float(self.Frame_count))
        self.Obj_avgarea = self.round(self.Obj_avgarea/float(self.Frame_count) / 1000.)

        self.Acc = -1 if gt_len <= 0 else self.round(gt_acc/float(gt_len) * 100)

    def round(self, val):
        return round(val,2)
    
    def print_info(self):
        # print(f"AI : {self.AI_count}, Trk : {self.Trk_count}, ratio : {self.Trk_AI_ratio}")
        # print(f"AI_time : {self.AI_time}, Trk_time : {self.Trk_time}")
        # print(f"FPS = {self.FPS:.2f}")
        print(f"ACC = {self.Acc}")

    def save_to_excel(self, file_name, res_acc, yolo_model):
        execl_path = "./auto_test_res.xlsx"
        # execl_path = "./auto_test_res_" + yolo_model + ".xlsx"
        if isfile(execl_path):
            workbook = load_workbook(execl_path)
        else:
            workbook = Workbook()

        sheet_exist = False
        sheet = workbook.active

        for shh in workbook.worksheets:
            if shh.title == "DIVA_1":
                sheet = shh
                sheet_exist = True
                break
            
        if not sheet_exist:
            sheet = workbook.create_sheet(title="DIVA_1")
            data_name = ['yolo_model', 'name', 'acc', 'avg_area', 'obj_count', 'max_obj']
            for i, value in enumerate(data_name):
                sheet.cell(1,i+1,value)

        last_row = sheet.max_row+1

        #task_case = "DIVA" if self.use_DIVA else "AI"
        # data_res = [file_name, self.Width, self.AI_count, self.AI_time, self.Trk_count, self.Trk_time, self.Trk_AI_ratio,
        #             self.Obj_avgnum, self.Obj_maxnum, self.Obj_avgarea, self.round(self.FPS), self.Acc]
        data_res = [yolo_model, file_name, res_acc, self.Obj_avgarea, self.Obj_avgnum, self.Obj_maxnum]
        for i, value in enumerate(data_res):
            #sheet.cell(last_row,i+1).number_format = numbers.FORMAT_NUMBER_00
            sheet.cell(last_row,i+1,value)  

        # Save the workbook
        workbook.save(execl_path)

class DIVA_mv_type(object):
    def __init__(self, diva_opt, _is_async=True):
        self.MVb = []
        self.cap = []
        self.video_name = diva_opt["stream_path"][0]
        self.use_diva = diva_opt["use_diva"]
        self.width = diva_opt["width"]
        self.height = diva_opt["height"]
        self.is_async = False if not self.use_diva else _is_async
        print(f"{self.is_async}........")

    def cap_init(self):
        pass

    def cap_read(self):
        pass

    def mv_init(self, width, height):
        pass

    def mv_get(self, frame, mv_all):
        pass

    def mv_release(self):
        pass

    @staticmethod
    def create(diva_opt, type_name='NVOF'):
        if type_name == 'FFmpeg':
            return MV_FFmpeg(diva_opt, False)
        elif type_name == 'NVOF':
            return MV_NVOF(diva_opt, True)
        elif type_name == 'FFOF':
            return MV_FFOF(diva_opt, False)
        else:
            raise ValueError("Unknown type specified")

class MV_FFmpeg(DIVA_mv_type):
    def cap_init(self):
        from c_pack.sdk.mvel import ffmpegcap
        self.cap = ffmpegcap()
        ret, self.width, self.height = self.cap.open(self.video_name)
        return self.width, self.height
    
    def cap_read(self):
        ret, frame, mv_all = self.cap.read()
        return ret, frame, mv_all
    
    def mv_init(self, width, height):
        pass

    def mv_get(self, frame, mv_all = [[],0,0]):
        return frame, mv_all

    def mv_release(self):
        self.cap.release()

class MV_NVOF(DIVA_mv_type):
    def cap_init(self):
        self.cap = cv2.VideoCapture(self.video_name)
        return self.width, self.height
    
    def cap_read(self):
        ret, frame = self.cap.read()
        return ret, frame, [[],0,0]
    
    def mv_init(self, width, height):
        if self.use_diva:
            import c_pack.sdk.appofcuda as appofcuda
            # self.MVb = appofcuda.appofcuda_get_mv(width, height, self.is_async, 16)
            w = int(width/16)
            h = int(height/16)
            print("w h", width, height, w, h)
            self.MVb = appofcuda.appofcuda_get_mv(width, height, True, 4*4, resize_w=w, resize_h=h)

    def mv_get(self, frame, mv_all = [[],0,0]):
        if self.use_diva:
            frame, mv_all = self.MVb.getnvof(frame)
        return frame, mv_all

    def mv_release(self):
        self.cap.release()
        if self.use_diva:
            self.MVb.release()

class MV_FFOF(DIVA_mv_type):
    def cap_init(self):
        self.cap = cv2.VideoCapture(self.video_name)
        return self.width, self.height
    
    def cap_read(self):
        ret, frame = self.cap.read()
        return ret, frame, [[],0,0]
    
    def mv_init(self, width, height):
        if self.use_diva:
            #import c_pack.sdk.appofcuda as appofcuda
            # from lib.cuof_resize import FlowCuda as cudaof
            # from lib.cuof_origin import FlowCuda as cudaof
            from lib.cuof import FlowCuda as cudaof
            #self.MVb = appofcuda.appofcuda_get_mv(width, height, self.is_async, 4*4)
            self.MVb = cudaof(width, height, size=16)

    def mv_get(self, frame, mv_all = [[],0,0]):
        if self.use_diva:
            # frame, mv_all = self.MVb.getnvof(frame)
            frame, mv_all = self.MVb.getcuof(frame)
            mv_all = (np.array(mv_all, dtype=np.int32),0,0)
            # print("mv_all:", mv_all)
        return frame, mv_all

    def mv_release(self):
        self.cap.release()
        # if self.use_diva:
        #     self.MVb.release()



        